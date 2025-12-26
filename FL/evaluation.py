import gc
from typing import Iterable, Optional, Sequence, Tuple

import numpy as np
import pandas as pd
import tensorflow as tf
from sklearn.metrics import (
    classification_report,
    confusion_matrix,
    f1_score,
    precision_score,
    recall_score,
)

PerClassMetrics = Tuple[np.ndarray, np.ndarray, np.ndarray]


def evaluate_model_streaming(
    model,
    X_path: str,
    y_path: str,
    num_classes: int,
    batch_size: int,
    is_gru: bool = False,
    chunk_size: int = 50000
):
    X_test = np.load(X_path, mmap_mode='r')
    y_test = np.load(y_path, mmap_mode='r')
    total_samples = len(X_test)

    all_y_true = []
    all_y_pred = []
    total_loss = 0.0
    total_observations = 0

    for chunk_start in range(0, total_samples, chunk_size):
        chunk_end = min(chunk_start + chunk_size, total_samples)
        X_chunk = np.array(X_test[chunk_start:chunk_end], dtype=np.float32)
        y_chunk = np.array(y_test[chunk_start:chunk_end])

        if is_gru and len(X_chunk.shape) == 2:
            X_chunk = np.expand_dims(X_chunk, axis=1)

        if len(y_chunk.shape) == 1 or y_chunk.shape[1] == 1:
            y_chunk_cat = tf.keras.utils.to_categorical(y_chunk, num_classes).astype(np.float32)
        else:
            y_chunk_cat = y_chunk.astype(np.float32)

        chunk_dataset = tf.data.Dataset.from_tensor_slices((X_chunk, y_chunk_cat)).batch(batch_size)
        chunk_loss = model.evaluate(chunk_dataset, verbose=0)[0]
        total_loss += chunk_loss * (chunk_end - chunk_start)
        total_observations += (chunk_end - chunk_start)

        chunk_preds = model.predict(chunk_dataset, verbose=0)
        chunk_y_pred = np.argmax(chunk_preds, axis=1)

        if len(y_chunk.shape) == 1:
            chunk_y_true = y_chunk
        else:
            chunk_y_true = np.argmax(y_chunk, axis=1)

        all_y_true.extend(chunk_y_true)
        all_y_pred.extend(chunk_y_pred)

        del X_chunk, y_chunk, y_chunk_cat, chunk_dataset, chunk_preds, chunk_y_pred, chunk_y_true
        gc.collect()

    y_true = np.array(all_y_true)
    y_pred = np.array(all_y_pred)

    avg_loss = total_loss / total_observations if total_observations > 0 else 0.0
    accuracy = np.mean(y_true == y_pred)
    f1_macro = f1_score(y_true, y_pred, average='macro', zero_division=0)
    precision_macro = precision_score(y_true, y_pred, average='macro', zero_division=0)
    recall_macro = recall_score(y_true, y_pred, average='macro', zero_division=0)

    f1_per_class = f1_score(y_true, y_pred, average=None, zero_division=0)
    precision_per_class = precision_score(y_true, y_pred, average=None, zero_division=0)
    recall_per_class = recall_score(y_true, y_pred, average=None, zero_division=0)
    cm = confusion_matrix(y_true, y_pred)

    del X_test, y_test, all_y_true, all_y_pred
    gc.collect()

    return avg_loss, accuracy, f1_macro, precision_macro, recall_macro, (f1_per_class, precision_per_class, recall_per_class), cm, y_true, y_pred


def evaluate_model_with_metrics(
    model,
    test_dataset: Iterable,
    num_classes: int,
    class_names: Optional[Sequence[str]] = None,
    round_num: Optional[int] = None,
    strategy_name: Optional[str] = None,
    partition_type: Optional[str] = None
):
    y_true = []
    y_pred = []
    total_loss = 0.0
    total_samples = 0

    base_model = model.base_model if hasattr(model, 'base_model') else model
    if hasattr(base_model, 'model'):
        base_model = base_model.model

    for batch_x, batch_y in test_dataset:
        batch_preds = base_model.predict_on_batch(batch_x)
        batch_y_pred = np.argmax(batch_preds, axis=1)
        y_pred.extend(batch_y_pred)

        if len(batch_y.shape) > 1 and batch_y.shape[1] > 1:
            batch_y_true = np.argmax(batch_y.numpy(), axis=1)
        else:
            batch_y_true = batch_y.numpy().astype(int)
        y_true.extend(batch_y_true)

        batch_loss = tf.keras.losses.categorical_crossentropy(batch_y, batch_preds).numpy().mean()
        total_loss += batch_loss * len(batch_y_true)
        total_samples += len(batch_y_true)

        del batch_preds, batch_y_pred, batch_y_true

    y_true = np.array(y_true)
    y_pred = np.array(y_pred)
    test_loss = total_loss / total_samples if total_samples > 0 else 0.0
    accuracy = np.mean(y_true == y_pred)

    f1_macro = f1_score(y_true, y_pred, average='macro', zero_division=0)
    precision_macro = precision_score(y_true, y_pred, average='macro', zero_division=0)
    recall_macro = recall_score(y_true, y_pred, average='macro', zero_division=0)

    f1_per_class = f1_score(y_true, y_pred, average=None, zero_division=0)
    precision_per_class = precision_score(y_true, y_pred, average=None, zero_division=0)
    recall_per_class = recall_score(y_true, y_pred, average=None, zero_division=0)
    cm = confusion_matrix(y_true, y_pred)

    if class_names is None:
        class_names = [f"Class_{i}" for i in range(num_classes)]

    class_report = classification_report(
        y_true,
        y_pred,
        target_names=class_names,
        zero_division=0
    )

    del y_pred, y_true
    gc.collect()

    return (
        test_loss,
        accuracy,
        f1_macro,
        precision_macro,
        recall_macro,
        (f1_per_class, precision_per_class, recall_per_class),
        cm,
        class_report,
    )


def create_enhanced_excel_report(
    excel_filename: str,
    main_results_df: pd.DataFrame,
    per_class_metrics: Optional[PerClassMetrics],
    class_names: Sequence[str],
    current_round: int,
    conf_matrix: Optional[np.ndarray]
) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows

    try:
        wb = openpyxl.load_workbook(excel_filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

    if 'Overall_Metrics' in wb.sheetnames:
        wb.remove(wb['Overall_Metrics'])
    ws_main = wb.create_sheet('Overall_Metrics')

    for row in dataframe_to_rows(main_results_df, index=False, header=True):
        ws_main.append(row)

    for cell in ws_main[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    if per_class_metrics is not None and isinstance(per_class_metrics, tuple) and len(per_class_metrics) == 3:
        sheet_name = f'Round_{current_round}_PerClass'
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        ws_per_class = wb.create_sheet(sheet_name)

        f1_per_class, precision_per_class, recall_per_class = per_class_metrics
        per_class_df = pd.DataFrame({
            'Class': class_names,
            'F1_Score': f1_per_class,
            'Precision': precision_per_class,
            'Recall': recall_per_class
        })

        for row in dataframe_to_rows(per_class_df, index=False, header=True):
            ws_per_class.append(row)

    wb.save(excel_filename)
