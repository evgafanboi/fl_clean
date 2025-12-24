import os
import argparse
import numpy as np
import tensorflow as tf
import gc
import logging
import pandas as pd
import time
import pickle
from datetime import datetime
from strategy.FedAvg import FedAvg
from strategy.FedProx import FedProx
from strategy.FedDyn import FedDyn
from strategy.FedCoMed import FedCoMed
from strategy.robust_filter import RobustFilterWeights
from models import dense, fedprox_wrapper, feddyn_wrapper
from sklearn.metrics import f1_score, precision_score, recall_score, confusion_matrix, classification_report
# Control GPU memory usage
gpus = tf.config.experimental.list_physical_devices('GPU')
if gpus:
    try:
        tf.config.experimental.set_memory_growth(gpus[0], True)
        gpu_options = tf.compat.v1.GPUOptions(
            per_process_gpu_memory_fraction=0.9,
            allow_growth=True,
            polling_active_delay=10,
            allocator_type='BFC'
        )
        config = tf.compat.v1.ConfigProto(gpu_options=gpu_options)
        config.gpu_options.allow_growth = True
        tf.compat.v1.keras.backend.set_session(tf.compat.v1.Session(config=config))
    except Exception as e:
        print(f"GPU configuration error: {e}")

class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    PURPLE = '\033[95m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'

def log_timestamp(logger, message):
    """Log message with timestamp"""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    logger.info(f"[{timestamp}] {message}")
    print(f"{bcolors.OKCYAN}[{timestamp}] {message}{bcolors.ENDC}")

def setup_logger(n_clients, partition_type, strategy="FedUp"):
    log_filename = f"results/{strategy}_{n_clients}client_{partition_type}.log"
    logging.basicConfig(
        filename=log_filename,
        level=logging.INFO,
        format='%(levelname)s: %(message)s',
        filemode='w'
    )
    print(bcolors.OKCYAN + f"Logging to {log_filename}" + bcolors.ENDC)
    
    # Setup detailed class metrics logger
    detailed_log_filename = log_filename.replace('.log', '_detailed_class_metrics.log')
    detailed_logger = logging.getLogger('detailed_metrics')
    detailed_logger.setLevel(logging.INFO)
    detailed_handler = logging.FileHandler(detailed_log_filename, mode='w')
    detailed_handler.setFormatter(logging.Formatter('%(message)s'))
    detailed_logger.addHandler(detailed_handler)
    detailed_logger.propagate = False
    
    return logging.getLogger(), log_filename, detailed_logger

def parse_partition_type(partition_type):
    if '-' in partition_type:
        parts = partition_type.split('-')
        if len(parts) == 2:
            return parts[0], int(parts[1])
    raise ValueError(f"Invalid partition type '{partition_type}'. Use format 'iid-10'")

def setup_paths(client_id, partition_type, client_count):
    partitions_root = os.path.join("data", "partitions")
    client_folder = f"{client_count}_client"
    partition_dir = os.path.join(partitions_root, client_folder, partition_type)
    
    paths = {
        'train_X': os.path.join(partition_dir, f"client_{client_id}_X_train.npy"),
        'train_y': os.path.join(partition_dir, f"client_{client_id}_y_train.npy"),
        'public_X': os.path.join(partition_dir, f"client_{client_id}_X_public.npy"),
        'public_y': os.path.join(partition_dir, f"client_{client_id}_y_public.npy"),
        'test_X': os.path.join("data", "X_test.npy"),
        'test_y': os.path.join("data", "y_test.npy")
    }
    
    # Check for local test sets
    test_X_path = os.path.join(partition_dir, f"client_{client_id}_X_test.npy")
    test_y_path = os.path.join(partition_dir, f"client_{client_id}_y_test.npy")
    
    if os.path.exists(test_X_path) and os.path.exists(test_y_path):
        paths['local_test_X'] = test_X_path
        paths['local_test_y'] = test_y_path
    
    return paths

def restore_full_partition(paths):
    """Restore original partition by concatenating public and private splits"""
    X_train = np.load(paths['train_X'], mmap_mode='r')
    y_train = np.load(paths['train_y'], mmap_mode='r')
    
    if os.path.exists(paths['public_X']):
        X_public = np.load(paths['public_X'])
        y_public = np.load(paths['public_y'])
        
        X_full = np.concatenate([np.array(X_train), X_public], axis=0)
        y_full = np.concatenate([np.array(y_train), y_public], axis=0)
        
        # Save restored partition
        np.save(paths['train_X'], X_full)
        np.save(paths['train_y'], y_full)
        
        return True
    
    return False

def create_model(input_dim, num_classes, batch_size=8192, strategy=None, client_id=None):
    if strategy and hasattr(strategy, 'name'):
        if strategy.name == "FedProx":
            return fedprox_wrapper.create_fedprox_dense_model(input_dim, num_classes, batch_size, strategy)
        elif strategy.name == "FedDyn":
            return feddyn_wrapper.create_feddyn_dense_model(input_dim, num_classes, batch_size, strategy, client_id)
    return dense.create_dense_model(input_dim, num_classes, batch_size)

def create_client_dataset(X_path, y_path, input_dim, num_classes, batch_size):
    def generator():
        X_mmap = np.load(X_path, mmap_mode='r')
        y_mmap = np.load(y_path, mmap_mode='r')
        total_samples = X_mmap.shape[0]
        chunk_size = 10000
        
        for start_idx in range(0, total_samples, chunk_size):
            end_idx = min(start_idx + chunk_size, total_samples)
            X_chunk = np.array(X_mmap[start_idx:end_idx], dtype=np.float32)
            y_chunk = np.array(y_mmap[start_idx:end_idx], dtype=np.float32)
            if len(y_chunk.shape) == 1 or y_chunk.shape[1] == 1:
                y_chunk = tf.keras.utils.to_categorical(y_chunk.astype(np.int32), num_classes=num_classes).astype(np.float32)
            yield X_chunk, y_chunk
            del X_chunk, y_chunk
    
    output_signature = (
        tf.TensorSpec(shape=(None, input_dim), dtype=tf.float32),
        tf.TensorSpec(shape=(None, num_classes), dtype=tf.float32)
    )
    dataset = tf.data.Dataset.from_generator(generator, output_signature=output_signature)
    return dataset.unbatch().batch(batch_size).prefetch(tf.data.AUTOTUNE)

_test_dataset_cache = None

def load_test_dataset(batch_size, num_classes):
    global _test_dataset_cache
    
    if _test_dataset_cache is None:
        X_test = np.load("data/X_test.npy")
        y_test = np.load("data/y_test.npy")
        X_test = np.asarray(X_test, dtype=np.float32)
        y_test = np.asarray(y_test, dtype=np.float32)
        if len(y_test.shape) == 1:
            y_test = tf.keras.utils.to_categorical(y_test.astype(np.int32), num_classes).astype(np.float32)
        _test_dataset_cache = (X_test, y_test)
        print(f"{bcolors.OKGREEN}Test dataset cached ({X_test.shape[0]} samples){bcolors.ENDC}")
    
    X_test, y_test = _test_dataset_cache
    dataset = tf.data.Dataset.from_tensor_slices((X_test, y_test))
    return dataset.batch(batch_size).prefetch(tf.data.AUTOTUNE)

def aggressive_memory_cleanup():
    tf.keras.backend.clear_session()
    tf.compat.v1.reset_default_graph()
    gc.collect()
    gc.collect()
    gc.collect()
    time.sleep(0.1)

def evaluate_model_streaming(model, X_path, y_path, num_classes, batch_size, is_gru=False, chunk_size=50000):
    """
    Evaluate model by streaming test data in chunks to minimize memory usage.
    Processes chunk_size samples at a time, aggregates metrics, then cleans up.
    """
    # Load full arrays with memory mapping (read-only, no RAM copy)
    X_test = np.load(X_path, mmap_mode='r')
    y_test = np.load(y_path, mmap_mode='r')
    total_samples = len(X_test)
    
    all_y_true = []
    all_y_pred = []
    total_loss = 0.0
    num_batches = 0
    
    # Process in chunks
    for chunk_start in range(0, total_samples, chunk_size):
        chunk_end = min(chunk_start + chunk_size, total_samples)
        
        # Load chunk into RAM (controlled size)
        X_chunk = np.array(X_test[chunk_start:chunk_end], dtype=np.float32)
        y_chunk = np.array(y_test[chunk_start:chunk_end])
        
        # Preprocess chunk
        if is_gru and len(X_chunk.shape) == 2:
            X_chunk = np.expand_dims(X_chunk, axis=1)
        
        if len(y_chunk.shape) == 1 or y_chunk.shape[1] == 1:
            y_chunk_cat = tf.keras.utils.to_categorical(y_chunk, num_classes).astype(np.float32)
        else:
            y_chunk_cat = y_chunk.astype(np.float32)
        
        # Create temporary dataset for this chunk
        chunk_dataset = tf.data.Dataset.from_tensor_slices((X_chunk, y_chunk_cat))
        chunk_dataset = chunk_dataset.batch(batch_size)
        
        # Evaluate chunk
        chunk_loss = model.evaluate(chunk_dataset, verbose=0)[0]
        total_loss += chunk_loss * (chunk_end - chunk_start)
        num_batches += (chunk_end - chunk_start)
        
        # Get predictions for chunk
        chunk_preds = model.predict(chunk_dataset, verbose=0)
        chunk_y_pred = np.argmax(chunk_preds, axis=1)
        
        # Get true labels
        if len(y_chunk.shape) == 1:
            chunk_y_true = y_chunk
        else:
            chunk_y_true = np.argmax(y_chunk, axis=1)
        
        # Accumulate
        all_y_true.extend(chunk_y_true)
        all_y_pred.extend(chunk_y_pred)
        
        # Aggressive cleanup after each chunk
        del X_chunk, y_chunk, y_chunk_cat, chunk_dataset, chunk_preds, chunk_y_pred, chunk_y_true
        gc.collect()
    
    # Convert to arrays
    y_true = np.array(all_y_true)
    y_pred = np.array(all_y_pred)
    
    # Calculate metrics
    avg_loss = total_loss / num_batches if num_batches > 0 else 0.0
    accuracy = np.mean(y_true == y_pred)
    
    f1_macro = f1_score(y_true, y_pred, average='macro', zero_division=0)
    precision_macro = precision_score(y_true, y_pred, average='macro', zero_division=0)
    recall_macro = recall_score(y_true, y_pred, average='macro', zero_division=0)
    
    f1_per_class = f1_score(y_true, y_pred, average=None, zero_division=0)
    precision_per_class = precision_score(y_true, y_pred, average=None, zero_division=0)
    recall_per_class = recall_score(y_true, y_pred, average=None, zero_division=0)
    cm = confusion_matrix(y_true, y_pred)
    
    per_class_metrics = (f1_per_class, precision_per_class, recall_per_class)
    
    # Cleanup
    del X_test, y_test, all_y_true, all_y_pred
    gc.collect()
    
    return avg_loss, accuracy, f1_macro, precision_macro, recall_macro, per_class_metrics, cm, y_true, y_pred

def evaluate_model_with_metrics(model, test_dataset, num_classes, class_names=None, round_num=None, strategy_name=None, partition_type=None):
    # Single-pass evaluation: get predictions AND labels together
    y_true = []
    y_pred = []
    total_loss = 0.0
    total_samples = 0
    
    # Handle model wrappers (FedProx, FedDyn) by accessing underlying model
    base_model = model.base_model if hasattr(model, 'base_model') else model
    if hasattr(base_model, 'model'):
        base_model = base_model.model
    
    for batch_x, batch_y in test_dataset:
        # Get predictions for this batch
        batch_preds = base_model.predict_on_batch(batch_x)
        batch_y_pred = np.argmax(batch_preds, axis=1)
        y_pred.extend(batch_y_pred)
        
        # Extract true labels
        if len(batch_y.shape) > 1 and batch_y.shape[1] > 1:
            batch_y_true = np.argmax(batch_y.numpy(), axis=1)
        else:
            batch_y_true = batch_y.numpy().astype(int)
        y_true.extend(batch_y_true)
        
        # Compute loss for this batch
        batch_loss = tf.keras.losses.categorical_crossentropy(batch_y, batch_preds).numpy().mean()
        total_loss += batch_loss * len(batch_y_true)
        total_samples += len(batch_y_true)
        
        del batch_preds, batch_y_pred, batch_y_true
    
    y_true = np.array(y_true)
    y_pred = np.array(y_pred)
    test_loss = total_loss / total_samples if total_samples > 0 else 0.0
    accuracy = np.mean(y_true == y_pred)
    
    # Calculate macro averages (better for multiclass)
    f1_macro = f1_score(y_true, y_pred, average='macro', zero_division=0)
    precision_macro = precision_score(y_true, y_pred, average='macro', zero_division=0)
    recall_macro = recall_score(y_true, y_pred, average='macro', zero_division=0)
    
    # Per-class metrics and confusion matrix
    f1_per_class = f1_score(y_true, y_pred, average=None, zero_division=0)
    precision_per_class = precision_score(y_true, y_pred, average=None, zero_division=0)
    recall_per_class = recall_score(y_true, y_pred, average=None, zero_division=0)
    cm = confusion_matrix(y_true, y_pred)
    
    if class_names is None:
        class_names = [f"Class_{i}" for i in range(num_classes)]
    
    class_report = classification_report(y_true, y_pred, target_names=class_names, zero_division=0)
    
    # Clean up prediction tensors
    del y_pred, y_true
    gc.collect()
    
    return (test_loss, accuracy, f1_macro, precision_macro, recall_macro, 
            (f1_per_class, precision_per_class, recall_per_class), cm)

def create_enhanced_excel_report(excel_filename, main_results_df, per_class_metrics, 
                               class_names, current_round, conf_matrix):
    import openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, PatternFill
    
    try:
        wb = openpyxl.load_workbook(excel_filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
    
    # Main results sheet
    if 'Overall_Metrics' in wb.sheetnames:
        wb.remove(wb['Overall_Metrics'])
    ws_main = wb.create_sheet('Overall_Metrics')
    
    for r in dataframe_to_rows(main_results_df, index=False, header=True):
        ws_main.append(r)
    
    for cell in ws_main[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Per-class metrics sheet
    if per_class_metrics is not None and isinstance(per_class_metrics, tuple) and len(per_class_metrics) == 3:
        sheet_name = f'Round_{current_round}_PerClass'
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        ws_per_class = wb.create_sheet(sheet_name)
        
        f1_per_class, precision_per_class, recall_per_class = per_class_metrics
        per_class_df = pd.DataFrame({
            'Class': class_names,
            'F1_Score': f1_per_class,
            'Precision': precision_per_class,
            'Recall': recall_per_class
        })
        
        for r in dataframe_to_rows(per_class_df, index=False, header=True):
            ws_per_class.append(r)
    
    wb.save(excel_filename)

def main():
    parser = argparse.ArgumentParser(description="Federated Learning Simulation")
    parser.add_argument("--n_clients", type=int, default=10)
    parser.add_argument("--partition_type", type=str, default="iid-10")
    parser.add_argument("--rounds", type=int, default=10)
    parser.add_argument("--strategy", type=str, default="FedAvg", help="Strategy: FedAvg, FedProx, FedDyn, FedCoMed, or RobustFilter")
    parser.add_argument("--feddyn_alpha", type=float, default=0.1, help="FedDyn alpha (paper shows best at 0.1)")
    parser.add_argument("--batch_size", type=int, default=8192)
    parser.add_argument("--epochs", type=int, default=5)
    parser.add_argument("--weights_cache_dir", type=str, default="temp_weights")
    parser.add_argument("--mu", type=float, default=0.01, help="FedProx proximal term")
    
    args = parser.parse_args()
    
    os.makedirs(args.weights_cache_dir, exist_ok=True)
    
    if args.strategy == "FedProx":
        strategy = FedProx(mu=args.mu)
        aggregator = strategy
        print(f"{bcolors.PURPLE}Using FedProx (mu={args.mu}){bcolors.ENDC}")
    elif args.strategy == "FedDyn":
        strategy = FedDyn(alpha=args.feddyn_alpha)
        aggregator = strategy
        print(f"{bcolors.PURPLE}Using FedDyn (alpha={args.feddyn_alpha}){bcolors.ENDC}")
    elif args.strategy == "FedCoMed":
        strategy = FedAvg()
        aggregator = FedCoMed()
        print(f"{bcolors.PURPLE}Using FedCoMed (Byzantine-robust coordinate-wise median){bcolors.ENDC}")
    elif args.strategy == "RobustFilter":
        strategy = FedAvg()
        aggregator = RobustFilterWeights(epsilon=0.2)
        print(f"{bcolors.PURPLE}Using RobustFilter (Byzantine-robust FedAvg, epsilon=0.2){bcolors.ENDC}")
    else:
        strategy = FedAvg()
        aggregator = strategy
        print(f"{bcolors.PURPLE}Using FedAvg{bcolors.ENDC}")
    
    partition_type, client_count = parse_partition_type(args.partition_type)
    n_clients = min(args.n_clients, client_count)
    
    logger, log_filename, detailed_logger = setup_logger(n_clients, partition_type, strategy=args.strategy)
    excel_filename = log_filename.replace('.log', '.xlsx')
    results_df = pd.DataFrame(columns=['Round', 'Loss', 'Accuracy', 'F1_Score', 'Precision', 'Recall'])
    
    class_names_file = os.path.join("data", "partitions", f"{client_count}_client", partition_type, "label_classes.npy")
    if os.path.exists(class_names_file):
        class_names = np.load(class_names_file, allow_pickle=True)
        class_names = [str(name) for name in class_names]
        num_classes = len(class_names)
    else:
        y_test = np.load(os.path.join("data", "y_test.npy"), mmap_mode='r')
        num_classes = len(np.unique(y_test))
        class_names = [f"Class_{i}" for i in range(num_classes)]
        del y_test
    
    sample_X = np.load(os.path.join("data", "X_test.npy"), mmap_mode='r')
    input_dim = sample_X.shape[1]
    del sample_X
    
    # Setup paths for all clients
    paths_list = []
    for i in range(n_clients):
        paths = setup_paths(str(i), partition_type, client_count)
        paths_list.append(paths)
    
    print(f"{bcolors.OKCYAN}Restoring full partitions...{bcolors.ENDC}")
    restored_count = sum(restore_full_partition(paths) for paths in paths_list)
    if restored_count > 0:
        print(f"{bcolors.OKGREEN}Restored {restored_count} partitions{bcolors.ENDC}")
    
    latest_weights = None
    pipeline_start_time = time.time()
    log_timestamp(logger, "=== FL PIPELINE STARTED ===")
    log_timestamp(logger, f"Strategy: {args.strategy}, Clients: {n_clients}, Rounds: {args.rounds}")
    round_times = []
    
    # Main federated learning loop
    for round_num in range(1, args.rounds + 1):
        round_start_time = time.time()
        
        logger.info(f"Round {round_num}/{args.rounds}")
        print(f"\n{bcolors.HEADER}Round {round_num}/{args.rounds}{bcolors.ENDC}")
        log_timestamp(logger, f"--- Round {round_num} started ---")
        
        weights_files = []
        sample_sizes = []
        client_losses = []
        participating_clients = []
        
        for i in range(n_clients):
            tf.keras.backend.clear_session()
            
            client_id = str(i)
            print(f"\n{bcolors.BOLD}Client {client_id}{bcolors.ENDC}")
            
            client_start_time = time.time()
            log_timestamp(logger, f"Client {client_id} training started")
            
            model = create_model(input_dim, num_classes, args.batch_size, strategy, i)
            
            if latest_weights is not None:
                model.set_weights(latest_weights)
            
            paths = paths_list[i]
            train_dataset = create_client_dataset(
                paths['train_X'], paths['train_y'], 
                input_dim, num_classes, args.batch_size
            )
            
            X_train_mmap = np.load(paths['train_X'], mmap_mode='r')
            X_train_size = X_train_mmap.shape[0]
            sample_sizes.append(X_train_size)
            del X_train_mmap
            
            print(f"Training client {client_id} for {args.epochs} epochs")
            history = model.fit(
                train_dataset,
                epochs=args.epochs,
                callbacks=[
                    tf.keras.callbacks.EarlyStopping(
                        monitor='loss', patience=3, restore_best_weights=True, verbose=1
                    )
                ],
                verbose=1
            )
            
            client_end_time = time.time()
            client_train_time = client_end_time - client_start_time
            log_timestamp(logger, f"Client {client_id} completed in {client_train_time:.2f}s")
            
            loss = float(history.history["loss"][-1]) if 'loss' in history.history else 0.0
            client_losses.append(loss)
            
            weights_file = os.path.join(args.weights_cache_dir, f"client_{client_id}_round_{round_num}_weights.pkl")
            
            if args.strategy == "FedDyn":
                feddyn_data = model.get_feddyn_update()
                with open(weights_file, 'wb') as f:
                    pickle.dump(feddyn_data, f)
            else:
                with open(weights_file, 'wb') as f:
                    pickle.dump(model.get_weights(), f)
            
            weights_files.append(weights_file)
            participating_clients.append(i)
            
            logger.info(f"Client {client_id}: Loss={loss:.4f}")
            print(f"{bcolors.WARNING}Client {client_id}: Loss={loss:.4f}{bcolors.ENDC}")
            
            del model, train_dataset, history
            aggressive_memory_cleanup()
        
        if not participating_clients:
            break
        
        weights_list = []
        for weights_file in weights_files:
            with open(weights_file, 'rb') as f:
                saved_data = pickle.load(f)
                if isinstance(saved_data, dict) and 'weights' in saved_data:
                    weights_list.append(saved_data['weights'])
                else:
                    weights_list.append(saved_data)

        print(f"\n{bcolors.OKBLUE}Aggregating weights ({args.strategy}){bcolors.ENDC}")
        
        if hasattr(strategy, 'name') and strategy.name == "FedDyn":
            aggregated_weights = strategy.aggregate(weights_list, sample_sizes, participating_clients)
        else:
            aggregated_weights = aggregator.aggregate(weights_list, sample_sizes)
        
        latest_weights = aggregated_weights
        del weights_list
        gc.collect()
        
        aggressive_memory_cleanup()
        eval_start_time = time.time()
        
        tf.keras.backend.clear_session()
        eval_model = create_model(input_dim, num_classes, args.batch_size, strategy)
        eval_model.set_weights(aggregated_weights)
        
        eval_batch_size = min(args.batch_size, 2048)
        test_dataset = load_test_dataset(eval_batch_size, num_classes)
        
        if round_num == args.rounds:
            print(f"{bcolors.HEADER}FINAL GLOBAL EVALUATION{bcolors.ENDC}")
        else:
            print(f"\n{bcolors.OKCYAN}Evaluating global model{bcolors.ENDC}")
        
        result = evaluate_model_with_metrics(
            eval_model, test_dataset, num_classes, class_names, 
            round_num, args.strategy, partition_type
        )
        test_loss, accuracy, f1, precision, recall = result[:5]
        per_class_metrics = result[5] if len(result) > 5 else None
        confusion_mat = result[6] if len(result) > 6 else None
        
        print(f"{bcolors.OKGREEN}[GLOBAL] Acc={accuracy:.4f}, Loss={test_loss:.4f}, F1={f1:.4f}{bcolors.ENDC}")
        
        eval_time = time.time() - eval_start_time
        log_timestamp(logger, f"Global eval completed in {eval_time:.2f}s")
        
        del eval_model, test_dataset
        aggressive_memory_cleanup()
        
        # # PERSONALIZED EVALUATION CODE REMOVED FOR COMPACTNESS
        
        new_row = pd.DataFrame({
            'Round': [round_num],
            'Loss': [test_loss],
            'Accuracy': [accuracy],
            'F1_Score': [f1],
            'Precision': [precision],
            'Recall': [recall]
        })
        results_df = pd.concat([results_df, new_row], ignore_index=True)
        
        # Save Excel only once per round, with per-class metrics only on final round
        if round_num == args.rounds and per_class_metrics is not None:
            create_enhanced_excel_report(excel_filename, results_df, per_class_metrics, 
                                       class_names, round_num, confusion_mat)
        else:
            results_df.to_excel(excel_filename, index=False)
        
        round_avg_loss = sum(client_losses) / len(client_losses) if client_losses else 0.0
        
        logger.info(f"Round {round_num} - Loss: {test_loss:.4f}, Acc: {accuracy:.4f}, F1: {f1:.4f}")
        print(f"{bcolors.OKGREEN}Round {round_num} completed - Loss: {test_loss:.4f}, Acc: {accuracy:.4f}, F1: {f1:.4f}{bcolors.ENDC}")
        
        round_time = time.time() - round_start_time
        round_times.append(round_time)
        log_timestamp(logger, f"--- Round {round_num} completed in {round_time:.2f}s ---")
        
        # Clean up weights files
        for weights_file in weights_files:
            if round_num != args.rounds:
                os.remove(weights_file)
        
        del aggregated_weights
        aggressive_memory_cleanup()
        time.sleep(1)
    
    # TIMESTAMP: Pipeline end with statistics
    pipeline_end_time = time.time()
    total_time = pipeline_end_time - pipeline_start_time
    avg_round_time = sum(round_times) / len(round_times) if round_times else 0
    
    log_timestamp(logger, f"All rounds completed")
    logger.info(f"Total time: {total_time:.2f} seconds ({total_time/60:.2f} minutes)")
    logger.info(f"Average time per round: {avg_round_time:.2f} seconds")
    logger.info(f"Total rounds completed: {len(round_times)}")
    
    print(f"{bcolors.OKGREEN}Pipeline completed!{bcolors.ENDC}")
    print(f"{bcolors.OKGREEN}Total time: {total_time:.2f} seconds ({total_time/60:.2f} minutes){bcolors.ENDC}")
    print(f"{bcolors.OKGREEN}Average time per round: {avg_round_time:.2f} seconds{bcolors.ENDC}")
    
    print(f"{bcolors.OKCYAN}Results saved to {excel_filename}{bcolors.ENDC}")

if __name__ == "__main__":
    main()